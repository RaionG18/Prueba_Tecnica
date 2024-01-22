import locale
import os
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
import requests
import time
import unidecode
import winreg

from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
from openpyxl.utils.cell import coordinate_from_string as cfs
from openpyxl.utils.cell import get_column_letter as gcl
from urllib.parse import urljoin

def getListSeparator():
    '''Retrieves the Windows list separator character from the registry'''
    aReg = winreg.ConnectRegistry(None, winreg.HKEY_CURRENT_USER)
    aKey = winreg.OpenKey(aReg, r"Control Panel\International")
    return winreg.QueryValueEx(aKey, "sList")[0]

def manejar_error(excepcion):
    print(f'Ocurrió un error: {excepcion}')

def convert_rng_to_df(tlc, l_col, l_row, sheet):
    first_col = cfs(tlc)[0]
    first_row = cfs(tlc)[1]

    rng = f"{first_col}{first_row+1}:{l_col}{l_row}"

    data_rows = []
    for row in sheet[rng]:
        data_rows.append([cell.value for cell in row])
    
    return pd.DataFrame(data_rows[2:], columns=data_rows[0])

def generar_dataframe(fecha_inicio, fecha_final):
    # Convertir las cadenas de fecha a objetos datetime
    fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d %H:%M:%S')
    fecha_final = datetime.strptime(fecha_final, '%Y-%m-%d %H:%M:%S')

    # Crear un rango de fechas intermedias con intervalo de 1 hora
    delta = timedelta(hours=1)  # Intervalo de 1 hora
    fechas_intermedias = [fecha_inicio + i * delta for i in range(int((fecha_final - fecha_inicio) / delta)+1)]

    return pd.DataFrame({'fecha_hora': fechas_intermedias})

def descargar_excel(url, nombre_archivo, fecha_actual):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Lanza una excepción para códigos de estado no exitosos

        # Carga el archivo Excel con openpyxl
        wb = load_workbook(BytesIO(response.content))
        
        # Especifica la hoja con el nombre "LDM"
        sheet_name = 'LDM'
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'No se encontró la hoja con el nombre "{sheet_name}" en el archivo Excel.')

        sheet = wb[sheet_name]

        section_headers = ['DEMANDA MÍNIMA', 'DEMANDA MEDIA', 'DEMANDA MÁXIMA']

        last_col = ''
        last_row = ''
        df_dict = {}  # Dictionary to hold the dataframes
        for cell in sheet['A']:  # Looping Column A only
            if cell.value in section_headers:
                tblname = cell.value  # Header of the Data Set found
                tlc = cell.coordinate  # Top Left Cell of the range
                start_row = cfs(tlc)[1]  #
                for x in range(1, sheet.max_column+1):  # Find the last used column for the data in this section
                     if cell.offset(row=1, column=x).value is None:
                        last_col = gcl(x)
                        break
                for y in range(1, sheet.max_row):  # Find the last used row for the data in this section
                    if cell.offset(row=y, column=1).value is None:
                        last_row = (start_row + y) - 1
                        if last_row != (start_row + 1):  
                            break

                # print(f"Range to convert for '{tblname}' is: '{tlc}:{last_col}{last_row}'")
                df_dict[tblname] = convert_rng_to_df(tlc, last_col, last_row, sheet)  # Convert to dataframe
                df_dict[tblname]["Planta Generadora"] = df_dict[tblname]["Planta Generadora"].apply(unidecode.unidecode)
                df_dict[tblname] = df_dict[tblname][df_dict[tblname]["Nemo"] == "JEN-C"]
                df_dict[tblname]["Banda"] = unidecode.unidecode(tblname)

                tblname_normalized = unidecode.unidecode(tblname)

                banda_mapping = {
                     "DEMANDA MINIMA": (("00:00:00", "05:00:00"), ("22:00:00", "23:00:00")),
                     "DEMANDA MEDIA": (("06:00:00", "17:00:00"),),
                     "DEMANDA MAXIMA": (("18:00:00", "21:00:00"),)
                    }

                tblname_normalized = unidecode.unidecode(tblname)

                # Crear un DataFrame vacío
                df_resultante = pd.DataFrame()
                if tblname_normalized in banda_mapping:
                    horas_intervalos = banda_mapping[tblname_normalized]

                    for hora_inicio, hora_final in horas_intervalos:
                        # Generar el DataFrame
                        df_temporal = generar_dataframe(f'{fecha_actual} {hora_inicio}', f'{fecha_actual} {hora_final}')
                        df_resultante = pd.concat([df_resultante, df_temporal], ignore_index=True)

                    df_resultante["Banda"] = tblname_normalized
                    df_dict[tblname] = pd.merge(df_dict[tblname], df_resultante, on='Banda')
                else:
                    pass

        wb.save(f"{nombre_archivo}")
        print("\n")
        print(f'Archivo "{nombre_archivo}" guardado exitosamente.')
        ### Print the DataFrames
        
        df_resultado = pd.DataFrame()
        for table_name, df in df_dict.items():
            df_resultado = pd.concat([df_resultado, df], ignore_index=True)

        df_resultado.columns = df.columns.str.strip()
        df_resultado = df_resultado.rename(columns={'Costo en US$/MWH': 'Costo'})
        df_resultado = df_resultado[['fecha_hora', 'Nemo', 'Planta Generadora', 'Potencia Disponible', 'Costo', 'FPNE', 'Banda']]
        df_resultado = df_resultado.sort_values(by='fecha_hora')

        return df_resultado
            
    except requests.exceptions.HTTPError as errh:
        manejar_error(errh)
    except requests.exceptions.ConnectionError as errc:
        manejar_error(errc)
    except requests.exceptions.Timeout as errt:
        manejar_error(errt)
    except requests.exceptions.RequestException as err:
        manejar_error(err)
    except pd.errors.EmptyDataError:
        print(f'No se encontró la hoja con el nombre "LDM" en el archivo Excel.')
    except Exception as e:
        manejar_error(e)
    
